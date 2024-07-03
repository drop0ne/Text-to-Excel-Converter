import tkinter as tk
from tkinter import filedialog, scrolledtext, ttk
import os
import subprocess
import sys
import time
import openpyxl
from openpyxl.styles import Font
import ctypes

# Hide the console window
ctypes.windll.user32.ShowWindow(ctypes.windll.kernel32.GetConsoleWindow(), 0)

class ProgressBar:
    def __init__(self, parent, total_steps, label):
        self.total_steps = total_steps
        self.current_step = 0
        self.label = tk.Label(parent, text=label)
        self.label.pack()
        self.progress = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(parent, variable=self.progress, maximum=total_steps)
        self.progress_bar.pack(fill=tk.X, padx=10, pady=5)

    def update(self):
        self.current_step += 1
        self.progress.set(self.current_step)
        self.label.update_idletasks()

    def finish(self):
        self.current_step = self.total_steps
        self.progress.set(self.current_step)
        self.label.update_idletasks()

class TextInputDialog:
    def __init__(self, parent):
        self.dialog_window = tk.Toplevel(parent)
        self.dialog_window.title("UTF-8 Text Input")
        self.text_area = scrolledtext.ScrolledText(self.dialog_window, wrap=tk.WORD, width=100, height=40)
        self.text_area.pack(padx=10, pady=10)
        self.submit_button = tk.Button(self.dialog_window, text="Submit", command=self.on_submit)
        self.submit_button.pack(pady=5)
        self.input_text = None

    def on_submit(self):
        self.input_text = self.text_area.get("1.0", tk.END)
        self.dialog_window.destroy()

    def get_input_text(self):
        self.dialog_window.grab_set()
        self.dialog_window.wait_window()
        return self.input_text

class FileDialog:
    @staticmethod
    def get_save_path():
        root = tk.Tk()
        root.withdraw()
        file_path = filedialog.asksaveasfilename(
            initialdir=r"D:\Users\Main Profile\Documents\Developer Files\My Documents",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        root.destroy()
        return file_path

class ExplorerOpener:
    @staticmethod
    def open_directory():
        path = r"D:\Users\Main Profile\Documents\Developer Files\My Documents"
        subprocess.run(f'explorer {path}', shell=True)

class Document:
    def __init__(self, text):
        self.text = text
        self.sections = []
        self.parse_text()

    def parse_text(self):
        sections = self.text.split('### ')
        for section_text in sections[1:]:
            self.sections.append(Section(section_text.strip()))

    def to_excel(self, file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Magnetic Bearings"
        row = 1

        for section in self.sections:
            row = section.write_to_excel(ws, row)

        wb.save(file_path)

class Section:
    def __init__(self, text):
        self.title = ""
        self.subsections = []
        self.parse_text(text)

    def parse_text(self, text):
        lines = text.split('\n')
        self.title = lines[0].strip()
        subsection_text = ""
        for line in lines[1:]:
            if line.startswith('#### '):
                if subsection_text:
                    self.subsections.append(SubSection(subsection_text.strip()))
                    subsection_text = ""
                subsection_text += line + '\n'
            else:
                subsection_text += line + '\n'
        if subsection_text:
            self.subsections.append(SubSection(subsection_text.strip()))

    def write_to_excel(self, ws, row):
        ws.append([self.title])
        ws[row][0].font = Font(bold=True, size=14)
        row += 1
        for subsection in self.subsections:
            row = subsection.write_to_excel(ws, row)
        return row

class SubSection:
    def __init__(self, text):
        self.title = ""
        self.content = []
        self.parse_text(text)

    def parse_text(self, text):
        lines = text.split('\n')
        self.title = lines[0].strip()
        content = ""
        for line in lines[1:]:
            if line.startswith('1. ') or line.startswith('- '):
                if content:
                    self.content.append(content.strip())
                    content = ""
            content += line + '\n'
        if content:
            self.content.append(content.strip())

    def write_to_excel(self, ws, row):
        ws.append([self.title])
        ws[row][0].font = Font(bold=True, size=12)
        row += 1
        for content in self.content:
            ws.append([content])
            row += 1
        return row

class Application:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Website to Excel Downloader")

        self.progress_frame = tk.Frame(self.root)
        self.progress_frame.pack(pady=10)

        self.input_text = None
        self.file_path = None

    def run(self):
        self.input_text = self.get_text_input()
        if not self.input_text.strip():
            self.quit_application()
            return

        self.show_progress_bar("Processing")
        self.file_path = FileDialog.get_save_path()
        if self.file_path:
            self.process_document()
            ExplorerOpener.open_directory()
            self.show_progress_bar("Finalizing")
        else:
            print("Save operation cancelled.")
        
        self.quit_application()

    def get_text_input(self):
        text_input_dialog = TextInputDialog(self.root)
        return text_input_dialog.get_input_text()

    def show_progress_bar(self, label):
        progress_bar = ProgressBar(self.progress_frame, total_steps=10, label=label)
        for _ in range(10):
            time.sleep(0.1)  # Simulate processing
            progress_bar.update()
        progress_bar.finish()

    def process_document(self):
        document = Document(self.input_text)
        document.to_excel(self.file_path)

    def quit_application(self):
        self.root.quit()
        self.root.destroy()

def main():
    app = Application()
    app.run()

if __name__ == "__main__":
    main()
